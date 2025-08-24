import openpyxl
import csv
from dataclasses import dataclass, asdict


@dataclass
class BaseRecord:
    M: int
    N: int
    K: int


@dataclass
class Record(BaseRecord):
    bias: bool = False
    dtype: str = "torch.bfloat16"
    outdtype: str = "torch.bfloat16"
    scaleAB: bool = False


def to_record(
    mnkRecord: BaseRecord,
    bias=False,
    dtype="torch.bfloat16",
    outdtype="torch.bfloat16",
    scaleAB=False,
) -> Record:
    return Record(
        M=mnkRecord.M,
        N=mnkRecord.N,
        K=mnkRecord.K,
        bias=bias,
        dtype=dtype,
        outdtype=outdtype,
        scaleAB=scaleAB,
    )


def excel_to_struct_list(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' 不存在！")
    sheet = wb[sheet_name]

    baseRecords = []

    # read D,E,F column (M, N, K) from 6 row
    for row in sheet.iter_rows(min_row=6, min_col=4, max_col=6, values_only=True):
        if all(v is None for v in row):
            continue
        M, N, K = row
        mnkrecord = BaseRecord(M=M, N=N, K=K)
        baseRecords.append(mnkrecord)

    return baseRecords


def save_structs_to_csv(records, csv_file, fieldnames):
    with open(csv_file, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow(asdict(rec))

    print(f"data write to {csv_file} success!!")


if __name__ == "__main__":
    excel_file = "MI355_OOTB_Kernel_Performance.xlsx"
    sheet_name = "Qwen32B-Gemm"
    csv_file_name = "qwen32B_untuned_gemm"

    mnkrecords = excel_to_struct_list(excel_file, sheet_name)
    records = [to_record(mnk) for mnk in mnkrecords]

    save_structs_to_csv(mnkrecords, f"{csv_file_name}.csv", ["M", "N", "K"])
    save_structs_to_csv(
        records,
        f"{csv_file_name}_bf16.csv",
        ["M", "N", "K", "bias", "dtype", "outdtype", "scaleAB"],
    )
